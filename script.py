import paramiko
import os
from datetime import datetime
#import pytz
import json
import threading
import csv
import pandas as pd

INFO = "\033[34m[INFO]\033[0m"
DEBUG = "\033[33m[DEBUG]\033[0m"
WARNING = "\033[33m[WARNING]\033[0m"
ERROR = "\033[31m[ERROR]\033[0m"
SUCCESS = "\033[32m[SUCCESS]\033[0m"
INPUT = "\033[33m[INPUT]\033[0m"


def banner():
    title = """
    ======================================================================
     â–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•—â–ˆâ–ˆâ•—  â–ˆâ–ˆâ•—â–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•— â–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•—â–ˆâ–ˆâ•—  â–ˆâ–ˆâ•—â–ˆâ–ˆâ•—     â–ˆâ–ˆâ•—â–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•—â–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•—
    â–ˆâ–ˆâ•”â•â•â•â•â•â–ˆâ–ˆâ•‘  â–ˆâ–ˆâ•‘â–ˆâ–ˆâ•”â•â•â•â•â•â–ˆâ–ˆâ•”â•â•â•â•â•â–ˆâ–ˆâ•‘ â–ˆâ–ˆâ•”â•â–ˆâ–ˆâ•‘     â–ˆâ–ˆâ•‘â–ˆâ–ˆâ•”â•â•â•â•â•â•šâ•â•â–ˆâ–ˆâ•”â•â•â•
    â–ˆâ–ˆâ•‘     â–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•‘â–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•—  â–ˆâ–ˆâ•‘     â–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•”â• â–ˆâ–ˆâ•‘     â–ˆâ–ˆâ•‘â–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•—   â–ˆâ–ˆâ•‘   
    â–ˆâ–ˆâ•‘     â–ˆâ–ˆâ•”â•â•â–ˆâ–ˆâ•‘â–ˆâ–ˆâ•”â•â•â•  â–ˆâ–ˆâ•‘     â–ˆâ–ˆâ•”â•â–ˆâ–ˆâ•— â–ˆâ–ˆâ•‘     â–ˆâ–ˆâ•‘â•šâ•â•â•â•â–ˆâ–ˆâ•‘   â–ˆâ–ˆâ•‘   
    â•šâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•—â–ˆâ–ˆâ•‘  â–ˆâ–ˆâ•‘â–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•—â•šâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•—â–ˆâ–ˆâ•‘  â–ˆâ–ˆâ•—â–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•—â–ˆâ–ˆâ•‘â–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•‘   â–ˆâ–ˆâ•‘   
     â•šâ•â•â•â•â•â•â•šâ•â•  â•šâ•â•â•šâ•â•â•â•â•â•â• â•šâ•â•â•â•â•â•â•šâ•â•  â•šâ•â•â•šâ•â•â•â•â•â•â•â•šâ•â•â•šâ•â•â•â•â•â•â•   â•šâ•â•   
                                                                        
    â–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•—  â–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•— â–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•—  â–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•— â–ˆâ–ˆâ–ˆâ•—   â–ˆâ–ˆâ–ˆâ•—â–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•—                 
    â–ˆâ–ˆâ•”â•â•â–ˆâ–ˆâ•—â–ˆâ–ˆâ•”â•â•â–ˆâ–ˆâ•—â–ˆâ–ˆâ•”â•â•â–ˆâ–ˆâ•—â–ˆâ–ˆâ•”â•â•â–ˆâ–ˆâ•—â–ˆâ–ˆâ–ˆâ–ˆâ•— â–ˆâ–ˆâ–ˆâ–ˆâ•‘â–ˆâ–ˆâ•”â•â•â•â•â•                 
    â–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•”â•â–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•‘â–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•”â•â–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•‘â–ˆâ–ˆâ•”â–ˆâ–ˆâ–ˆâ–ˆâ•”â–ˆâ–ˆâ•‘â–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•—                 
    â–ˆâ–ˆâ•”â•â•â•â• â–ˆâ–ˆâ•”â•â•â–ˆâ–ˆâ•‘â–ˆâ–ˆâ•”â•â•â–ˆâ–ˆâ•—â–ˆâ–ˆâ•”â•â•â–ˆâ–ˆâ•‘â–ˆâ–ˆâ•‘â•šâ–ˆâ–ˆâ•”â•â–ˆâ–ˆâ•‘â•šâ•â•â•â•â–ˆâ–ˆâ•‘                 
    â–ˆâ–ˆâ•‘     â–ˆâ–ˆâ•‘  â–ˆâ–ˆâ•‘â–ˆâ–ˆâ•‘  â–ˆâ–ˆâ•‘â–ˆâ–ˆâ•‘  â–ˆâ–ˆâ•‘â–ˆâ–ˆâ•‘ â•šâ•â• â–ˆâ–ˆâ•‘â–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ–ˆâ•‘â–ˆâ–ˆâ•—              
    â•šâ•â•     â•šâ•â•  â•šâ•â•â•šâ•â•  â•šâ•â•â•šâ•â•  â•šâ•â•â•šâ•â•     â•šâ•â•â•šâ•â•â•â•â•â•â•â•šâ•â• 
    JUNIPER NETWORK
    Version: v2.0
    ======================================================================
    """
    print(title)
    
def input_jumphost():
    print("Please input your jumphost credentials:")
    jh_ip = input(f"{INPUT} JUMPSHOT IP: ").strip()
    jh_user = input(f"{INPUT} JUMPSHOT USER: ").strip()
    jh_pass = input(f"{INPUT} JUMPSHOT PASSWORD: ").strip()
    print("\n")
    return jh_ip, jh_user, jh_pass

def input_device_list():
    print("Please input your device list ip (comma-separated):")
    print("Example: 192.168.1.1,192.168.1.2")
    device_list = input(f"{INPUT} DEVICE LIST IP: ").strip().split(",")
    print("\n")
    print(f"{INFO} Please input device credentials:")
    print(f"{INFO} Default Username: appeventf")
    
    use_default = input("Use default credentials for all devices? (Y/n): ").strip().lower()
    if use_default in ['', 'y', 'yes']:
        user = "appeventf"
        pwd = "Penguin24!"
    else:
        user = input(f"{INPUT} DEVICE USERNAME: ").strip()
        pwd = input(f"{INPUT} DEVICE PASSWORD: ").strip()
    print("\n")
    # return device_list.split(",")
    return [(ip.strip(), user, pwd) for ip in device_list]

def input_command_list():
    print(f"{INFO} Please select input command list (comma-separated):")
    print("1). List Command Checklist Parameter")
    print("2). Input Command List Custom")
    
    option = input(f"{INPUT} Select Option (1/2): ").strip()
    print("\n")
    if option == "1":
        command_list = [
            "show interface detail | display json",
            "show vlan extensive | display json",
            "show ethernet-switching table | display json",
            "show evpn database | display json",
            "show configuration | display inheritance | display json",
            "show chassis alarm | display json",
            "show system alarm | display json",
            "show chassis routing-engine | display json",
            "show route summary table bgp.evpn.0 | display json",
            "show arp no-resolve | display json",
            "show bgp neighbor | display json",
            "show bfd session detail | display json",
            "show ddos-protection protocols | display json"
        ]
        print(f"{INFO} Using default command list:")
        for cmd in command_list:
            print(f"  > {cmd}")
        print()
    else:
        print(f"{INFO} Please input your command list (comma-separated):")
        print("Example: show interface terse, show vlan extensive")
        command_input = input(f"{INPUT} COMMAND LIST:\n").strip()
        print("\n")
        command_list = [cmd.strip() for cmd in command_input.split(",") if cmd.strip()]

    return command_list

def connect_jumphost(jh_ip, jh_user, jh_pass):
    jh_client = paramiko.SSHClient()
    jh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    jh_client.connect(jh_ip, username=jh_user, password=jh_pass)
    return jh_client

def run_on_target_via_jump(jh_client, target_ip, target_user, target_pass, command, timeout=10):
    try:
        # Open fresh channel to target (don't reuse)
        jump_transport = jh_client.get_transport()
        if not jump_transport:
            return "{ERROR}: Jump host transport not available"

        channel = jump_transport.open_channel(
            "direct-tcpip",
            (target_ip, 22),
            ("127.0.0.1", 0)
        )

        # Use Paramiko.Transport directly (fresh)
        target_transport = paramiko.Transport(channel)
        target_transport.start_client(timeout=timeout)
        target_transport.auth_password(username=target_user, password=target_pass)

        # Use Paramiko SSHClient on top of fresh transport
        target = paramiko.SSHClient()
        target.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        target._transport = target_transport

        stdin, stdout, stderr = target.exec_command(command, timeout=timeout)

        output = stdout.read().decode(errors="ignore").strip()
        error = stderr.read().decode(errors="ignore").strip()

        stdout.close()
        stderr.close()
        stdin.close()
        target.close()
        target_transport.close()

        if error:
            return f"{ERROR}: {error}"
        return output

    except paramiko.ssh_exception.AuthenticationException as e:
        return f"Authentication failed: {e}"
    except paramiko.ssh_exception.SSHException as e:
        return f"{ERROR} SSH: {e}"
    except Exception as e:
        return f"Exception: {e}"

def process_device(ip, user, pwd, jh_client, command_list, output_dir):
    print(f"\n[+] Thread: Connecting to {ip} via jumphost...")

    for cmd in command_list:
        print(f"  > [{ip}] Running: {cmd}")
        result = run_on_target_via_jump(jh_client, ip, user, pwd, cmd)

        # [INFO] OS windows melarang penggunaan PIPE, opsi Bersihkan karakter ilegal dari command
        safe_cmd = cmd.replace(" ", "_").replace("|", "_").replace("/", "_").replace("\\", "_")
        filename = f"{ip}-{safe_cmd}.txt"

        if "json" in cmd:
            try:
                json.loads(result)
                filename = f"{ip}-{safe_cmd}.json"
            except json.JSONDecodeError:
                print(f"  ! WARNING: Output from {ip} for '{cmd}' is not valid JSON, saving as .txt")

        with open(os.path.join(output_dir, filename), "w", encoding="utf-8", errors="replace") as f:
            f.write(result)


def load_device_list(filename):
    devices = []
    with open(filename, "r", encoding="utf-8") as f:
        for line in f:
            parts = line.strip().split("|||")
            if len(parts) == 3:
                ip, user, password = parts
                devices.append((ip, user, password))
    return devices

def load_command_list(filename):
    commands = []
    with open(filename, "r", encoding="utf-8") as f:
        for line in f:
            cmd = line.strip()
            if cmd:
                commands.append(cmd)
    return commands

def summary_rate(timestamp):
    json_folder = f"output-{timestamp}"
    output_file = f"{json_folder}/summary_rate-{timestamp}.csv"
    with open(output_file, "w", encoding="utf-8") as out:
        out.write("ip,interface,input_rate,output_rate\n")  # Header CSV
        for filename in os.listdir(json_folder):
            if "show_interface_detail" in filename and filename.endswith(".json"):
                ip = filename.split(".json")[0].split("-")[0]
                with open(os.path.join(json_folder, filename), "r", encoding="utf-8") as file:
                    data = json.load(file)

                try:
                    interfaces = data["interface-information"][0]["physical-interface"]
                    for interface in interfaces:
                        interface_name = interface["name"][0]["data"]
                        if any(prefix in interface_name for prefix in ["ge", "xe", "et", "ae", "irb"]):
                            if "irb" in interface_name:
                                for logical in interface.get("logical-interface", []):
                                    logical_name = logical["name"][0]["data"]
                                    input_rate = logical["traffic-statistics"][0]["input-bytes"][0]["data"]
                                    output_rate = logical["traffic-statistics"][0]["output-bytes"][0]["data"]
                                    line = f"{ip},{logical_name},{input_rate},{output_rate}\n"
                                    out.write(line)
                            else:
                                input_rate = interface["traffic-statistics"][0]["input-bps"][0]["data"]
                                output_rate = interface["traffic-statistics"][0]["output-bps"][0]["data"]
                                line = f"{ip},{interface_name},{input_rate},{output_rate}\n"
                                out.write(line)
                                for logical in interface.get("logical-interface", []):
                                    if "traffic-statistics" in logical:
                                        logical_name = logical["name"][0]["data"]
                                        input_rate = logical["traffic-statistics"][0]["input-bytes"][0]["data"]
                                        output_rate = logical["traffic-statistics"][0]["output-bytes"][0]["data"]
                                        line = f"{ip},{logical_name},{input_rate},{output_rate}\n"
                                        out.write(line)
                                    elif "lag-traffic-statistics" in logical:
                                        lag_stats = logical["lag-traffic-statistics"][0].get("lag-bundle", [{}])[0]
                                        logical_name = logical["name"][0]["data"]
                                        input_rate = lag_stats["input-bps"][0]["data"]
                                        output_rate = lag_stats["output-bps"][0]["data"]
                                        line = f"{ip},{logical_name},{input_rate},{output_rate}\n"
                                        out.write(line)
                                    else: 
                                        print(f"No traffic statistics for logical interface {logical['name'][0]['data']} on {ip}")
                except Exception as e:
                    print(f"{ERROR} processing {filename}: {e}")

def vlan_extensive(timestamp):
    json_folder = f"output-{timestamp}"
    output_file = f"{json_folder}/vlan_extensive-{timestamp}.csv"
    with open(output_file, "w", encoding="utf-8") as out:
        out.write("ip,vlan_name,vlan_state,vlan_tag,vlan_vxlan_enabled,tagged_interface,untagged_interface,mac_count\n")  # Header CSV
        with open(f"{json_folder}/vlan_summary-{timestamp}.csv", "w", encoding="utf-8") as txt_out:
            txt_out.write(f"ip,vlan_active_count,vlan_destroy_count\n") # Header for summary CSV
            for filename in os.listdir(json_folder):
                if "show_vlan_extensive" in filename and filename.endswith(".json"):
                    ip = filename.split(".json")[0].split("-")[0]
                    with open(os.path.join(json_folder, filename), "r", encoding="utf-8") as file:
                        data = json.load(file)

                    try:
                        vlans = data["l2ng-l2ald-vlan-instance-information"][0]["l2ng-l2ald-vlan-instance-group"]
                        vlan_active_count = 0
                        vlan_destroy_count = 0
                        for vlan in vlans:
                            vlan_name = vlan["l2ng-l2rtb-vlan-name"][0]["data"]
                            vlan_state = vlan["l2ng-l2rtb-instance-state"][0]["data"]
                            vlan_tag = vlan["l2ng-l2rtb-vlan-tag"][0]["data"]
                            vlan_vxlan_enabled = vlan["l2ng-l2rtb-vlan-vxlan-enabled"][0]["data"]
                            tagged_interfaces = vlan["l2ng-l2rtb-vlan-member-tagged-count"][0]["data"]
                            untagged_interfaces = vlan["l2ng-l2rtb-vlan-member-untagged-count"][0]["data"]
                            mac_count = vlan["l2ng-l2rtb-macs-learned"][0]["data"]
                            if vlan_state == "Active":
                                vlan_active_count += 1
                            else:
                                vlan_destroy_count += 1
                            line = f"{ip},{vlan_name},{vlan_state},{vlan_tag},{vlan_vxlan_enabled},{tagged_interfaces},{untagged_interfaces},{mac_count}\n"
                            out.write(line)
                        summary_line = f"{ip},{vlan_active_count},{vlan_destroy_count}\n"
                        txt_out.write(summary_line)
                    except Exception as e:
                        print(f"{ERROR} processing {filename}: {e}")

def ethernet_switching_table(timestamp):
    json_folder = f"output-{timestamp}"
    output_file = f"{json_folder}/ethernet_switching_table-{timestamp}.csv"
    with open(output_file, "w", encoding="utf-8") as out:
        out.write("ip,mac_addr,interface\n")  # Header CSV
        with open(f"{json_folder}/ethernet_switching_table_summary-{timestamp}.csv", "w", encoding="utf-8") as txt_out:
            txt_out.write(f"ip,mac_count\n") # Header for summary CSV
            for filename in os.listdir(json_folder):
                if "show_ethernet-switching_table" in filename and filename.endswith(".json"):
                    ip = filename.split(".json")[0].split("-")[0]
                    with open(os.path.join(json_folder, filename), "r", encoding="utf-8") as file:
                        data = json.load(file)

                    try:
                        all_mac = data["l2ng-l2ald-rtb-macdb"][0]["l2ng-l2ald-mac-entry-vlan"][0]["l2ng-mac-entry"]
                        for mac in all_mac:
                            mac_addr = mac["l2ng-l2-mac-address"][0]["data"]
                            interface = mac["l2ng-l2-mac-logical-interface"][0]["data"]
                            line = f"{ip},{mac_addr},{interface}\n"
                            out.write(line)
                        summary_line = f"{ip},{len(all_mac)}\n"
                        txt_out.write(summary_line)
                    except Exception as e:
                        print(f"{ERROR} processing {filename}: {e}")

def evpn_database(timestamp):
    json_folder = f"output-{timestamp}"
    output_file = f"{json_folder}/evpn_database-{timestamp}.csv"
    with open(output_file, "w", encoding="utf-8") as out:
        out.write("ip,vni_id,mac_addr,active_source,active_source_timestamp,ip_address\n")  # Header CSV
        with open(f"{json_folder}/evpn_database_summary-{timestamp}.csv", "w", encoding="utf-8") as txt_out:
            txt_out.write(f"ip,evpn_count\n") # Header for summary CSV
            for filename in os.listdir(json_folder):
                if "show_evpn_database" in filename and filename.endswith(".json"):
                    ip = filename.split(".json")[0].split("-")[0]
                    with open(os.path.join(json_folder, filename), "r", encoding="utf-8") as file:
                        data = json.load(file)

                    try:
                        all_evpn = data["evpn-database-information"][0]["evpn-database-instance"][0]["mac-entry"]
                        #print(ip, "has", len(all_evpn), "EVPN entries")
                        for evpn in all_evpn:
                            vni_id = evpn["vni-id"][0]["data"]
                            mac_addr = evpn["mac-address"][0]["data"]
                            active_source = evpn["active-source"][0]["data"]
                            active_source_timestamp = evpn["active-source-timestamp"][0]["data"]
                            ip_address = ""
                            if "ip-address" in evpn:
                                ip_address_list = [item["data"] for item in evpn["ip-address"] if "data" in item]
                                ip_address = ";".join(ip_address_list)
                            line = f"{ip},{vni_id},{mac_addr},{active_source},{active_source_timestamp},{ip_address}\n"
                            out.write(line)
                        summary_line = f"{ip},{len(all_evpn)}\n"
                        txt_out.write(summary_line)
                    except Exception as e:
                        print(f"{ERROR} processing {filename}: {e}")

def save_config_to_json(timestamp):
    json_folder = f"output-{timestamp}"
    for filename in os.listdir(json_folder):
        if "show_configuration" in filename:
            full_path = os.path.join(json_folder, filename)
            try:
                with open(full_path, "r", encoding="utf-8") as file:
                    data = json.load(file)

                hostname = data["configuration"]["system"]["host-name"]
                if "-RE" in hostname:
                    hostname = hostname.split("-RE")[0]

                output_dir = os.path.join(json_folder, "config")
                os.makedirs(output_dir, exist_ok=True)

                output_path = os.path.join(output_dir, f"{hostname}.json")
                with open(output_path, "w", encoding="utf-8") as out_file:
                    json.dump(data, out_file, indent=4)

            except json.JSONDecodeError:
                print(f"{WARNING}: {filename} is not valid JSON, skipping")
            except UnicodeDecodeError as e:
                print(f"{ERROR}: Cannot read {filename} due to encoding issue: {e}")
            except Exception as e:
                print(f"{ERROR}: Unexpected processing {filename}: {e}")


def get_esi(timestamp):
    json_folder = f"output-{timestamp}"
    output_file = f"{json_folder}/esi-lag-{timestamp}.csv"
    with open(output_file, "w", encoding="utf-8") as out:
        out.write("ip,interface,esi_id\n")  # Header CSV
        for filename in os.listdir(json_folder):
            if filename.endswith(".json"):
                ip = filename.split(".json")[0].split("-")[0]
                with open(os.path.join(json_folder, filename), "r", encoding="utf-8") as file:
                    data = json.load(file)

                try:
                    all_interface = data["configuration"]["interfaces"]["interface"]
                    #print(all_interface)
                    for interface in all_interface:
                        if "esi" in interface:
                            line = f"{ip},{interface["name"]},{interface["esi"]["identifier"]}\n"
                            out.write(line)
                except Exception as e:
                    print(f"{ERROR} processing {filename}: {e}")
    
def alarm(timestamp, type):
    json_folder = f"output-{timestamp}"
    output_file = f"{json_folder}/{type}_alarm-{timestamp}.csv"
    with open(output_file, "w", encoding="utf-8") as out:
        out.write(f"ip,{type}_alarm_count\n")  # Header CSV
        for filename in os.listdir(json_folder):
            if f"show_{type}_alarm.json" in filename:
                ip = filename.split(".json")[0].split("-")[0]
                with open(os.path.join(json_folder, filename), "r", encoding="utf-8") as file:
                    data = json.load(file)

                try:
                    alarm_summary = data["alarm-information"][0]["alarm-summary"][0]
                    alarm_count = ""
                    if "no-active-alarms" in alarm_summary:
                        alarm_count = 0
                    elif "active-alarm-count" in alarm_summary:
                        alarm_count = alarm_summary["active-alarm-count"][0]["data"]
                    line = f"{ip},{alarm_count}\n"
                    #line = f"{ip},{vlan_name},{vlan_state},{vlan_tag},{vlan_vxlan_enabled},{tagged_interfaces},{untagged_interfaces},{mac_count}\n"
                    out.write(line)
                except Exception as e:
                    print(f"{ERROR} processing {filename}: {e}")

def chassis_re(timestamp):
    json_folder = f"output-{timestamp}"
    output_file = f"{json_folder}/chassis_util-{timestamp}.csv"
    with open(output_file, "w", encoding="utf-8") as out:
        out.write("ip,cpu_usage,mem_usage\n")  # Header CSV
        for filename in os.listdir(json_folder):
            if "chassis_routing-engine" in filename and filename.endswith(".json"):
                ip = filename.split(".json")[0].split("-")[0]
                with open(os.path.join(json_folder, filename), "r", encoding="utf-8") as file:
                    data = json.load(file)

                try:
                    routing_engine = data["route-engine-information"][0]["route-engine"]
                    if len(routing_engine) > 1:
                        for index, re_slot in enumerate(routing_engine):
                            if re_slot["mastership-state"][0]["data"] == "master":
                                routing_engine = routing_engine[index]  
                                break
                    else:
                        routing_engine = routing_engine[0] 
                    cpu_usage = f"{100 - int(routing_engine["cpu-idle"][0]["data"])} percent"
                    mem_usage = f"{routing_engine["memory-buffer-utilization"][0]["data"]} percent"
                    line = f"{ip},{cpu_usage},{mem_usage}\n" 
                    out.write(line)
                except Exception as e:
                    print(f"{ERROR} processing {filename}: {e}")

def route_evpn_summary(timestamp):
    json_folder = f"output-{timestamp}"
    output_file = f"{json_folder}/route-evpn-summary-{timestamp}.csv"
    with open(output_file, "w", encoding="utf-8") as out:
        out.write("ip,destinations,routes\n")  # Header CSV
        for filename in os.listdir(json_folder):
            if "show_route_summary_table_bgp.evpn" in filename and filename.endswith(".json"):
                ip = filename.split(".json")[0].split("-")[0]
                with open(os.path.join(json_folder, filename), "r", encoding="utf-8") as file:
                    data = json.load(file)

                try:
                    route_summary = data["route-summary-information"][0]["route-table"][0]
                    destination_count = f"{route_summary["destination-count"][0]["data"]} Destinations"
                    route_count = f"{route_summary["total-route-count"][0]["data"]} Routes"
                    line = f"{ip},{destination_count},{route_count}\n" 
                    out.write(line)
                except Exception as e:
                    print(f"{ERROR} processing {filename}: {e}")

def arp(timestamp):
    json_folder = f"output-{timestamp}"
    output_file = f"{json_folder}/arp_database-{timestamp}.csv"
    with open(output_file, "w", encoding="utf-8") as out:
        out.write("ip,mac_addr,address,interface\n")  # Header CSV
        with open(f"{json_folder}/arp_summary-{timestamp}.csv", "w", encoding="utf-8") as txt_out:
            txt_out.write(f"ip,arp_count\n") # Header for summary CSV
            for filename in os.listdir(json_folder):
                if "show_arp" in filename and filename.endswith(".json"):
                    ip = filename.split(".json")[0].split("-")[0]
                    with open(os.path.join(json_folder, filename), "r", encoding="utf-8") as file:
                        data = json.load(file)

                    try:
                        all_arp = data["arp-table-information"][0]["arp-table-entry"]
                        for arp in all_arp:
                            mac_addr = arp["mac-address"][0]["data"]
                            address = arp["ip-address"][0]["data"]
                            interface = arp["interface-name"][0]["data"]
                            line = f"{ip},{mac_addr},{address},{interface}\n"
                            out.write(line)
                        summary_line = f"{ip},{len(all_arp)}\n"
                        txt_out.write(summary_line)
                    except Exception as e:
                        print(f"{ERROR} processing {filename}: {e}")
 
def bgp_summary(timestamp):
    json_folder = f"output-{timestamp}"
    output_file = f"{json_folder}/bgp_neighbor-{timestamp}.csv"
    with open(output_file, "w", encoding="utf-8") as out:
        out.write("ip,peer,asn,flap_count,last_event_flap,state\n")  # Header CSV
        with open(f"{json_folder}/bgp_summary-{timestamp}.csv", "w", encoding="utf-8") as txt_out:
            txt_out.write(f"ip,establish_count,down_count\n") # Header for summary CSV
            for filename in os.listdir(json_folder):
                if "show_bgp_neighbor" in filename and filename.endswith(".json"):
                    ip = filename.split(".json")[0].split("-")[0]
                    with open(os.path.join(json_folder, filename), "r", encoding="utf-8") as file:
                        data = json.load(file)

                    try:
                        all_bgp = data["bgp-information"][0]["bgp-peer"]
                        total_established = 0
                        total_down = 0
                        for bgp in all_bgp:
                            peer = bgp["peer-address"][0]["data"]
                            peer_asn = bgp["peer-as"][0]["data"]
                            flap_count = bgp["flap-count"][0]["data"]
                            if "last-flap-event" in bgp:
                                last_event_flap = bgp["last-flap-event"][0]["data"]
                            else:
                                last_event_flap = bgp["last-event"][0]["data"]
                            state = bgp["peer-state"][0]["data"]
                            if state == "Established":
                                total_established += 1
                            else:
                                total_down += 1
                            line = f"{ip},{peer},{peer_asn},{flap_count},{last_event_flap},{state}\n"
                            out.write(line)
                        summary_line = f"{ip},up : {total_established},down : {total_down}\n"
                        txt_out.write(summary_line)
                    except Exception as e:
                        print(f"{ERROR} processing {filename}: {e}")            

def bfd_summary(timestamp):
    json_folder = f"output-{timestamp}"
    output_file = f"{json_folder}/bfd_session-{timestamp}.csv"
    with open(output_file, "w", encoding="utf-8") as out:
        out.write("ip,peer,state,uptime\n")  # Header CSV
        with open(f"{json_folder}/bfd_summary-{timestamp}.csv", "w", encoding="utf-8") as txt_out:
            txt_out.write(f"ip,up_count,down_count\n") # Header for summary CSV
            for filename in os.listdir(json_folder):
                if "show_bfd_session_detail" in filename and filename.endswith(".json"):
                    ip = filename.split(".json")[0].split("-")[0]
                    with open(os.path.join(json_folder, filename), "r", encoding="utf-8") as file:
                        data = json.load(file)

                    try:
                        all_bfd = data["bfd-session-information"][0]["bfd-session"  ]
                        total_up = 0
                        total_down = 0
                        for bfd in all_bfd:
                            peer = bfd["session-neighbor"][0]["data"]
                            uptime = ""
                            state = bfd["session-state"][0]["data"]
                            if state == "Up":
                                total_up += 1
                                uptime = bfd["session-up-time"][0]["data"]
                            else:
                                total_down += 1
                            line = f"{ip},{peer},{state},{uptime}\n"
                            out.write(line)
                        summary_line = f"{ip},up : {total_up},down : {total_down}\n"
                        txt_out.write(summary_line)
                    except Exception as e:
                        print(f"{ERROR} processing {filename}: {e}")

def ddos_summary(timestamp):
    json_folder = f"output-{timestamp}"
    output_file = f"{json_folder}/ddos_summary-{timestamp}.csv"
    with open(output_file, "w", encoding="utf-8") as out:
        out.write("ip,total_packet,mod_packet,rcvd_packet,violation_packet\n")  # Header CSV
        for filename in os.listdir(json_folder):
            if "show_ddos-protection_protocols" in filename and filename.endswith(".json"):
                ip = filename.split(".json")[0].split("-")[0]
                with open(os.path.join(json_folder, filename), "r", encoding="utf-8") as file:
                    data = json.load(file)
                try:
                    all_ddos = data["ddos-protocols-information"][0]
                    total_packet_type = all_ddos["total-packet-types"][0]["data"]
                    mod_packet_types = all_ddos["mod-packet-types"][0]["data"]
                    rcvd_packet = all_ddos["packet-types-rcvd-packets"][0]["data"]
                    violation_packet = all_ddos["packet-types-in-violation"][0]["data"]
                    line = f"{ip},{total_packet_type},{mod_packet_types},{rcvd_packet},{violation_packet}\n"
                    out.write(line)
                except Exception as e:
                    print(f"{ERROR} processing {filename}: {e}")

def convert_arp_summary_to_json(timestamp):
    csv_folder = f"output-{timestamp}"
    # Ambil timestamp dari nama file
    output_file = f"{csv_folder}/arp_summary-{timestamp}.csv"
    db = {}
    with open(output_file, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            ip = row["ip"]
            arp_count = row["arp_count"]
            
            if ip not in db:
                db[ip] = {}
            db[ip][timestamp] = {
                "arp-entries": f"{arp_count} entries"
            }
    
    return db

def convert_vlan_summary_to_json(timestamp):
    csv_folder = f"output-{timestamp}"
    # Ambil timestamp dari nama file
    output_file = f"{csv_folder}/vlan_summary-{timestamp}.csv"
    db = {}
    with open(output_file, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            ip = row["ip"]
            vlan_active_count = row["vlan_active_count"]
            vlan_destroy_count = row["vlan_destroy_count"]

            if ip not in db:
                db[ip] = {}
            db[ip][timestamp] = {
                "vlan_summary": f"{vlan_active_count} active, {vlan_destroy_count} destroyed"
            }

    return db

def convert_system_alarm_to_json(timestamp):
    csv_folder = f"output-{timestamp}"
    # Ambil timestamp dari nama file
    output_file = f"{csv_folder}/system_alarm-{timestamp}.csv"
    db = {}
    with open(output_file, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            ip = row["ip"]
            system_alarm_count = row["system_alarm_count"]

            if ip not in db:
                db[ip] = {}
            db[ip][timestamp] = {
                "system_alarm": f"{system_alarm_count} active alarm"
            }

    return db

def convert_chassis_alarm_to_json(timestamp):
    csv_folder = f"output-{timestamp}"
    # Ambil timestamp dari nama file
    output_file = f"{csv_folder}/chassis_alarm-{timestamp}.csv"
    db = {}
    with open(output_file, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            ip = row["ip"]
            chassis_alarm_count = row["chassis_alarm_count"]

            if ip not in db:
                db[ip] = {}
            db[ip][timestamp] = {
                "chassis_alarm": f"{chassis_alarm_count} active alarm"
            }

    return db

def convert_route_evpn_summary_to_json(timestamp):
    csv_folder = f"output-{timestamp}"
    # CREDIT : FITRAHHHHH, CGPT, CLAUDE
    # Ambil timestamp dari nama file
    output_file = f"{csv_folder}/route-evpn-summary-{timestamp}.csv"
    db = {}
    with open(output_file, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            ip = row["ip"]
            destinations = row["destinations"]
            routes = row["routes"]

            if ip not in db:
                db[ip] = {}
            db[ip][timestamp] = {
                "route-evpn-summary": f"Destination : {destinations}, Routes : {routes}"
            }

    return db

def convert_evpn_database_summary_to_json(timestamp):
    csv_folder = f"output-{timestamp}"
    # Ambil timestamp dari nama file
    output_file = f"{csv_folder}/evpn_database_summary-{timestamp}.csv"
    db = {}
    with open(output_file, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            ip = row["ip"]
            evpn_count = row["evpn_count"]

            if ip not in db:
                db[ip] = {}
            db[ip][timestamp] = {
                "evpn-entries": f"{evpn_count} entries"
            }

    return db

def convert_ethernet_switching_table_summary_to_json(timestamp):
    csv_folder = f"output-{timestamp}"
    # Ambil timestamp dari nama file
    output_file = f"{csv_folder}/ethernet_switching_table_summary-{timestamp}.csv"
    db = {}
    with open(output_file, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            ip = row["ip"]
            mac_count = row["mac_count"]

            if ip not in db:
                db[ip] = {}
            db[ip][timestamp] = {
                "mac-entries": f"{mac_count} entries"
            }

    return db

def convert_ddos_summary_to_json(timestamp):
    csv_folder = f"output-{timestamp}"
    # Ambil timestamp dari nama file
    output_file = f"{csv_folder}/ddos_summary-{timestamp}.csv"
    db = {}
    with open(output_file, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            ip = row["ip"]
            total_packet = int(row["total_packet"])
            mod_packet = row["mod_packet"]
            rcvd_packet = row["rcvd_packet"]
            violation_packet = row["violation_packet"]

            if ip not in db:
                db[ip] = {}
            db[ip][timestamp] = {
                "ddos_summary": f"Total : {total_packet}, Mod : {mod_packet}, Rcvd : {rcvd_packet}, Violation : {violation_packet}"
            }

    return db

def convert_chassis_util_summary_to_json(timestamp):
    csv_folder = f"output-{timestamp}"
    # Ambil timestamp dari nama file
    output_file = f"{csv_folder}/chassis_util-{timestamp}.csv"
    db = {}
    with open(output_file, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            ip = row["ip"]
            cpu_usage = row["cpu_usage"]
            mem_usage = row["mem_usage"]

            if ip not in db:
                db[ip] = {}
            db[ip][timestamp] = {
                "chassis-util": f"CPU : {cpu_usage}, RAM : {mem_usage}"
            }

    return db

def convert_bfd_summary_to_json(timestamp):
    csv_folder = f"output-{timestamp}"
    # Ambil timestamp dari nama file
    output_file = f"{csv_folder}/bfd_summary-{timestamp}.csv"
    db = {}
    with open(output_file, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            ip = row["ip"]
            up_count = row["up_count"]
            down_count = row["down_count"]

            if ip not in db:
                db[ip] = {}
            db[ip][timestamp] = {
                "bfd-summary": f"{up_count}, {down_count}"
            }

    return db

def convert_bgp_summary_to_json(timestamp):
    csv_folder = f"output-{timestamp}"
    # Ambil timestamp dari nama file
    output_file = f"{csv_folder}/bgp_summary-{timestamp}.csv"
    db = {}
    with open(output_file, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            ip = row["ip"]
            establish_count = row["establish_count"]
            down_count = row["down_count"]

            if ip not in db:
                db[ip] = {}
            db[ip][timestamp] = {
                "bgp-summary": f"{establish_count}, {down_count}"
            }

    return db

def load_existing_db(file_path):
    if not os.path.exists(file_path):
        print("ðŸ“‚ File db.json belum ada. Membuat baru...")
        return {}
    
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read().strip()
            if not content:
                print("âš ï¸  File db.json kosong. Menginisialisasi data kosong...")
                return {}
            return json.loads(content)
    except json.JSONDecodeError as e:
        print(f"âŒ Gagal memuat db.json karena format tidak valid: {e}")
        print("âš ï¸  Menggunakan database kosong sementara.")
        return {}
    except Exception as e:
        print(f"âŒ Terjadi kesalahan saat membaca db.json: {e}")
        return {}

def save_db(file_path, db):
    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(db, f, indent=2)
    print(f"âœ… db.json berhasil diperbarui: {file_path}")

def merge_to_db(existing_db, new_data):
    for ip, ts_data in new_data.items():
        if ip not in existing_db:
            existing_db[ip] = {}

        for ts, params in ts_data.items():
            if ts not in existing_db[ip]:
                existing_db[ip][ts] = {}

            for key, value in params.items():
                # Kalau parameter belum ada atau ingin diupdate
                existing_db[ip][ts][key] = value

    return existing_db

# === MAIN ===
# Banner
banner()

# Load Credential
jh_ip, jh_user, jh_pass = input_jumphost()

# Command and IP List
command_list = input_command_list()
device_list = input_device_list()
# command_list=load_command_list("command.txt")
# device_list = load_device_list("device_cred.txt")

# Hanya konek ke jumphost sekali
jh_client = connect_jumphost(jh_ip, jh_user, jh_pass)

#jakarta = pytz.timezone("Asia/Jakarta")
timestamp = datetime.now().strftime("%y%m%d-%H%M")
#timestamp = datetime.now(jakarta).strftime("%y%m%d-%H%M")
output_dir = f"output-{timestamp}"
os.makedirs(output_dir, exist_ok=True)


threads = []
for ip, user, pwd in device_list:
    t = threading.Thread(
        target=process_device,
        args=(ip, user, pwd, jh_client, command_list, output_dir)
    )
    t.start()
    threads.append(t)

# Tunggu semua thread selesai
for t in threads:
    t.join()

# Tutup koneksi jumphost setelah semua selesai
jh_client.close()

checklist_param = 0
# Check if any command show interface detail | display json
if any("show interface detail" in cmd and "display json" in cmd for cmd in command_list):
    summary_rate(timestamp)
    
# Check if any command show vlan extensive | display json
if any("show vlan extensive" in cmd and "display json" in cmd for cmd in command_list):
    checklist_param += 1
    vlan_extensive(timestamp)

# Check if any command show ethernet-switching table | display json
if any("show ethernet-switching table" in cmd and "display json" in cmd for cmd in command_list):
    checklist_param += 1
    ethernet_switching_table(timestamp)

# Check if any command show ethernet-switching table | display json
if any("show evpn database" in cmd and "display json" in cmd for cmd in command_list):
    checklist_param += 1
    evpn_database(timestamp)

# Save configuration to JSON file
# Check if any command show configuration | display json | display inheritance
if any("show configuration" in cmd and "display json" in cmd for cmd in command_list):
    save_config_to_json(timestamp)

if any("show chassis alarm" in cmd and "display json" in cmd for cmd in command_list):
    checklist_param += 1
    alarm(timestamp, "chassis")
if any("show system alarm" in cmd and "display json" in cmd for cmd in command_list):
    checklist_param += 1
    alarm(timestamp, "system")

if any("show chassis routing-engine" in cmd and "display json" in cmd for cmd in command_list):
    checklist_param += 1
    chassis_re(timestamp)

if any("show route summary table bgp.evpn.0" in cmd and "display json" in cmd for cmd in command_list):
    checklist_param += 1
    route_evpn_summary(timestamp)
    
if any("show arp no-resolve" in cmd and "display json" in cmd for cmd in command_list):
    checklist_param += 1
    arp(timestamp)

if any("show bgp neighbor" in cmd and "display json" in cmd for cmd in command_list):
    checklist_param += 1
    bgp_summary(timestamp)

if any("show bfd session detail" in cmd and "display json" in cmd for cmd in command_list):
    checklist_param += 1
    bfd_summary(timestamp)

if any("show ddos-protection protocols" in cmd and "display json" in cmd for cmd in command_list):
    checklist_param += 1
    ddos_summary(timestamp)

if checklist_param == 11:
    print("âœ… Semua parameter berhasil dikumpulkan.")
    arp_data = convert_arp_summary_to_json(timestamp)
    vlan_data = convert_vlan_summary_to_json(timestamp)
    bgp_data = convert_bgp_summary_to_json(timestamp)
    bfd_data = convert_bfd_summary_to_json(timestamp)
    chassis_alarm_data = convert_chassis_alarm_to_json(timestamp)
    chassis_util_data = convert_chassis_util_summary_to_json(timestamp)
    ddos_data = convert_ddos_summary_to_json(timestamp)
    ethernet_swtiching_data = convert_ethernet_switching_table_summary_to_json(timestamp)
    system_alarm_data = convert_system_alarm_to_json(timestamp)
    route_vpn_data = convert_route_evpn_summary_to_json(timestamp)
    evpn_data = convert_evpn_database_summary_to_json(timestamp)

    db_path = "db.json"
    existing_db = load_existing_db(db_path)

    # Merge dua data secara bertingkat
    all_data = [
        arp_data,
        vlan_data,
        bgp_data,
        bfd_data,
        chassis_alarm_data,
        chassis_util_data,
        ddos_data,
        ethernet_swtiching_data,
        system_alarm_data,
        route_vpn_data,
        evpn_data
    ]

    for dataset in all_data:
        existing_db = merge_to_db(existing_db, dataset)

    # Simpan hasil
    save_db(db_path, existing_db)
    def export_db_to_csv(db_path, output_csv):
        with open(db_path, encoding="utf-8") as f:
            db = json.load(f)

        all_timestamps = set()
        all_keys = set()

        # Ambil semua timestamp dan key unik
        for ip, ts_data in db.items():
            all_timestamps.update(ts_data.keys())
            for ts in ts_data.values():
                all_keys.update(ts.keys())

        # Urutkan timestamp biar kolom rapih
        all_timestamps = sorted(all_timestamps)
        all_keys = sorted(all_keys)

        rows = []

        for ip, ts_data in db.items():
            for key in all_keys:
                row = [ip, key]
                for ts in all_timestamps:
                    value = ts_data.get(ts, {}).get(key, "")
                    row.append(value)
                rows.append(row)

        # Header
        header = ["IP", "checklist"] + all_timestamps

        # Tulis ke CSV
        with open(output_csv, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(header)
            writer.writerows(rows)

        print(f"âœ… CSV berhasil dibuat: {output_csv}")

    def export_db_to_excel(db_path, output_excel):
        import pandas as pd
        from datetime import datetime
        
        with open(db_path, encoding="utf-8") as f:
            db = json.load(f)

        all_timestamps = set()
        all_keys = set()

        # Ambil semua timestamp dan key unik
        for ip, ts_data in db.items():
            all_timestamps.update(ts_data.keys())
            for ts in ts_data.values():
                all_keys.update(ts.keys())

        all_timestamps = sorted(all_timestamps)
        all_keys = sorted(all_keys)

        # Buat mapping untuk nama yang lebih user-friendly
        key_mapping = {
            'arp-entries': 'ARP Entries',
            'vlan_summary': 'VLAN Summary',
            'bgp-summary': 'BGP Summary',
            'bfd-summary': 'BFD Summary',
            'chassis_alarm': 'Chassis Alarm',
            'chassis-util': 'Chassis Utilization',
            'ddos_summary': 'DDoS Summary',
            'ethernet-switching': 'MAC Entries',
            'system_alarm': 'System Alarm',
            'route-evpn-summary': 'Route EVPN Summary',
            'evpn-entries': 'EVPN Entries'
        }

        # Konversi timestamp ke format yang lebih readable
        def format_timestamp(ts):
            try:
                # Asumsi format: YYMMDD-HHMM
                date_part = ts.split('-')[0]
                time_part = ts.split('-')[1]
                year = '20' + date_part[:2]
                month = date_part[2:4]
                day = date_part[4:6]
                hour = time_part[:2]
                minute = time_part[2:4]
                dt = datetime(int(year), int(month), int(day), int(hour), int(minute))
                return dt.strftime('%Y-%m-%d %H:%M')
            except:
                return ts

        formatted_timestamps = [format_timestamp(ts) for ts in all_timestamps]

        # Buat struktur data untuk Excel dengan IP grouping
        rows = []
        for ip in sorted(db.keys()):
            ts_data = db[ip]
            
            # Tambahkan row untuk setiap metric dari IP ini
            for key in all_keys:
                friendly_key = key_mapping.get(key, key.replace('_', ' ').title())
                row = {"IP Address": ip, "Metric": friendly_key}
                
                # Tambahkan data untuk setiap timestamp
                for i, ts in enumerate(all_timestamps):
                    value = ts_data.get(ts, {}).get(key, "N/A")
                    row[formatted_timestamps[i]] = value
                
                rows.append(row)

        df = pd.DataFrame(rows)
        
        # Buat Excel dengan formatting
        with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Network Monitoring Report', index=False)
            
            # Dapatkan workbook dan worksheet
            workbook = writer.book
            worksheet = writer.sheets['Network Monitoring Report']
            
            # Import styling modules
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils import get_column_letter
            
            # Define colors dan styles
            header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=12)
            cell_font = Font(size=10)
            change_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light pink
            
            # Border style
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Format header row
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Format data rows dan apply merging + comparison logic
            current_ip = None
            ip_start_row = 2
            
            for row_num in range(2, len(df) + 2):
                ip_address = df.iloc[row_num - 2]['IP Address']
                
                # Check if we're starting a new IP group
                if current_ip != ip_address:
                    # Merge cells for previous IP if exists
                    if current_ip is not None and row_num > ip_start_row:
                        worksheet.merge_cells(f'A{ip_start_row}:A{row_num-1}')
                        # Center align the merged cell
                        worksheet.cell(row=ip_start_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
                    
                    current_ip = ip_address
                    ip_start_row = row_num
                
                # Format all cells in this row
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.font = cell_font
                    cell.border = thin_border
                    
                    # Center align IP dan Metric columns
                    if col_num <= 2:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Merge cells for the last IP group
            if current_ip is not None:
                worksheet.merge_cells(f'A{ip_start_row}:A{len(df)+1}')
                worksheet.cell(row=ip_start_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
            
            # Apply comparison logic and highlight changes
            if len(all_timestamps) > 1:
                timestamp_cols = list(range(3, len(df.columns) + 1))  # Columns with timestamp data
                
                for row_num in range(2, len(df) + 2):
                    # Compare consecutive timestamps
                    for i in range(len(timestamp_cols) - 1):
                        col_a = timestamp_cols[i]
                        col_b = timestamp_cols[i + 1]
                        
                        value_a = worksheet.cell(row=row_num, column=col_a).value
                        value_b = worksheet.cell(row=row_num, column=col_b).value
                        
                        # Convert to string for comparison, handle None/N/A values
                        str_a = str(value_a) if value_a is not None else "N/A"
                        str_b = str(value_b) if value_b is not None else "N/A"
                        
                        # If values are different, highlight the second column (col_b)
                        if str_a != str_b:
                            worksheet.cell(row=row_num, column=col_b).fill = change_fill
            
            # Auto-adjust column widths
            for col_num, column_title in enumerate(df.columns, 1):
                column_letter = get_column_letter(col_num)
                max_length = len(str(column_title))
                
                for row in range(2, len(df) + 2):
                    cell_value = worksheet.cell(row=row, column=col_num).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                
                # Set column width dengan minimum dan maximum limits
                adjusted_width = min(max(max_length + 2, 12), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze panes untuk header dan IP/Metric columns
            worksheet.freeze_panes = 'C2'

        print(f"âœ… Excel berhasil dibuat : {output_excel}")

    # Contoh pemanggilan
    export_db_to_csv("db.json", "report.csv")
    export_db_to_excel("db.json", "report.xlsx")
